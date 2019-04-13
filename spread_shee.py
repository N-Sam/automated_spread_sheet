import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#print(cell.value) to get the item in a cell at a particular coordinate
#print(first_sheet.max_row) to get the total number of rows in a sheet

def proces_workbook(file_name):
    my_wb = xl.load_workbook(file_name)
    first_sheet = my_wb['Sheet1']
    cell = first_sheet.cell(1, 1)

    for row in range(2, first_sheet.max_row+1): #i started from 2 to ignore headings and incremented 1 because range function is size -1
        cell = first_sheet.cell(row, 3)
        updated_marks = cell.value + 5
        updated_marks_cell = first_sheet.cell(row, 4)
        updated_marks_cell.value = updated_marks


    values = Reference(first_sheet, min_row=2, max_row=first_sheet.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    first_sheet.add_chart(chart, 'e2')


    my_wb.save(file_name) # here you will overwrite the existing file
